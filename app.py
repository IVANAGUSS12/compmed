from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import time
import os, io, json
import uuid
import threading
import re
import textwrap
import shutil
from glob import glob
from difflib import SequenceMatcher

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'medcontrol-cemic-2026'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(DATA_DIR, "medcontrol.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

PDF_JOBS = {}
PDF_JOBS_LOCK = threading.Lock()


def _set_pdf_job(job_id, **changes):
    with PDF_JOBS_LOCK:
        if job_id in PDF_JOBS:
            PDF_JOBS[job_id].update(changes)


def _buscar_pdf_indicaciones_paciente(pid):
    """Devuelve la mejor ruta disponible del PDF de indicaciones original del paciente."""
    preferido = os.path.join(app.config['UPLOAD_FOLDER'], f'paciente_{pid}_indicaciones.pdf')
    if os.path.exists(preferido):
        return preferido

    patrones = [
        os.path.join(app.config['UPLOAD_FOLDER'], f'tmp_{pid}_*indicaciones*.pdf'),
        os.path.join(app.config['UPLOAD_FOLDER'], f'tmp_{pid}*indicaciones*.pdf'),
        os.path.join(BASE_DIR, '*indicaciones*.pdf'),
    ]
    candidatos = []
    for p in patrones:
        candidatos.extend(glob(p))
    if not candidatos:
        return None
    candidatos.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidatos[0]


def _procesar_pdfs_job(pid, job_id, tmp_ind, tmp_fact, reemplazar):
    try:
        with app.app_context():
            from utils.pdf_parser import parse_indicaciones, parse_facturados

            t0 = time.time()
            _set_pdf_job(job_id, status='running', progress=5, message='Iniciando lectura rápida de PDFs...')

            # Fase 1: intento rapido sin OCR (texto nativo del PDF)
            administrados, indicados = parse_indicaciones(tmp_ind, use_ocr_fallback=False)
            _set_pdf_job(job_id, progress=20, message='PDF de indicaciones leído (modo rápido).')

            facturados = parse_facturados(tmp_fact, use_ocr_fallback=False)
            _set_pdf_job(job_id, progress=35, message='PDF de facturados leído (modo rápido).')

            # Fase 2: fallback OCR solo si no salio data en modo rapido
            if not administrados and not indicados:
                _set_pdf_job(job_id, message='Activando OCR en indicaciones...')

                def cb_ind(n, total):
                    prog = 35 + int((n / max(total, 1)) * 25)
                    _set_pdf_job(job_id, progress=min(prog, 60), message=f'OCR indicaciones: página {n}/{total}')

                administrados, indicados = parse_indicaciones(
                    tmp_ind,
                    callback=cb_ind,
                    use_ocr_fallback=True
                )

            if not facturados:
                _set_pdf_job(job_id, message='Activando OCR en facturados...')

                def cb_fact(n, total):
                    prog = 60 + int((n / max(total, 1)) * 20)
                    _set_pdf_job(job_id, progress=min(prog, 80), message=f'OCR facturados: página {n}/{total}')

                facturados = parse_facturados(
                    tmp_fact,
                    callback=cb_fact,
                    use_ocr_fallback=True
                )

            if not administrados and not indicados and not facturados:
                _set_pdf_job(job_id, status='error', progress=100, error='No se pudieron extraer datos de los PDFs.')
                return

            _set_pdf_job(job_id, progress=85, message='Guardando datos en base...')

            if reemplazar:
                Administrado.query.filter_by(paciente_id=pid).delete()
                Indicado.query.filter_by(paciente_id=pid).delete()
                Facturado.query.filter_by(paciente_id=pid).delete()
                db.session.flush()

            n_adm = 0
            n_ind = 0
            n_fact = 0

            for r in administrados:
                fecha = (r.get('fecha') or '').strip()
                nombre = (r.get('nombre_original') or '').strip()
                if not fecha or not nombre:
                    continue
                db.session.add(Administrado(
                    paciente_id=pid,
                    fecha=fecha,
                    hora_enfermero=(r.get('hora_enfermero') or '').strip(),
                    nombre_original=nombre,
                    estado='pendiente'
                ))
                n_adm += 1

            for r in indicados:
                fecha = (r.get('fecha') or '').strip()
                droga = (r.get('droga') or '').strip()
                if not fecha or not droga:
                    continue
                db.session.add(Indicado(
                    paciente_id=pid,
                    fecha=fecha,
                    medico=(r.get('medico') or '').strip(),
                    droga=droga,
                    dosis_via=(r.get('dosis_via') or '').strip(),
                    observacion=''
                ))
                n_ind += 1

            for r in facturados:
                fecha = (r.get('fecha') or '').strip()
                nombre = (r.get('nombre') or '').strip()
                if not fecha or not nombre:
                    continue
                try:
                    cantidad = int(r.get('cantidad', 1) or 1)
                except Exception:
                    cantidad = 1
                db.session.add(Facturado(
                    paciente_id=pid,
                    fecha=fecha,
                    nombre=nombre,
                    cantidad=cantidad,
                    codigo=(r.get('codigo') or '').strip(),
                    valor=(r.get('valor') or '').strip(),
                ))
                n_fact += 1

            db.session.commit()
            _set_pdf_job(job_id, progress=92, message='Aplicando equivalencias...')

            cambios_equiv = aplicar_equivalencias(pid)

            db.session.add(LogCambio(
                paciente_id=pid,
                entidad='importacion_pdf',
                entidad_id=0,
                campo='cargar_pdfs',
                valor_anterior='',
                valor_nuevo=json.dumps({
                    'administrados': n_adm,
                    'indicados': n_ind,
                    'facturados': n_fact,
                    'equivalencias_aplicadas': cambios_equiv,
                    'reemplazar': reemplazar
                }, ensure_ascii=False),
                usuario='usuario',
                descripcion='Importacion automatica desde 2 PDFs'
            ))
            db.session.commit()

            _set_pdf_job(
                job_id,
                status='done',
                progress=100,
                message=f'Listo en {time.time() - t0:.1f}s',
                result={
                    'administrados': n_adm,
                    'indicados': n_ind,
                    'facturados': n_fact,
                    'equivalencias_aplicadas': cambios_equiv,
                    'reemplazar': reemplazar,
                }
            )
    except Exception as e:
        db.session.rollback()
        _set_pdf_job(job_id, status='error', progress=100, error=f'Error al procesar PDFs: {e}')
    finally:
        for p in (tmp_ind, tmp_fact):
            if os.path.exists(p):
                os.remove(p)

# ──────────────────────────────────────────
# MODELOS
# ──────────────────────────────────────────

class Paciente(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    nombre        = db.Column(db.String(200), nullable=False)
    hc            = db.Column(db.String(50))
    dni           = db.Column(db.String(20))
    plan          = db.Column(db.String(50))
    cama          = db.Column(db.String(20))
    fecha_ingreso = db.Column(db.String(20))
    fecha_egreso  = db.Column(db.String(20))
    notas         = db.Column(db.Text)
    creado_en     = db.Column(db.DateTime, default=datetime.utcnow)

    administrados = db.relationship('Administrado', backref='paciente', lazy=True, cascade='all,delete')
    indicados     = db.relationship('Indicado',     backref='paciente', lazy=True, cascade='all,delete')
    facturados    = db.relationship('Facturado',    backref='paciente', lazy=True, cascade='all,delete')
    cambios       = db.relationship('LogCambio',    backref='paciente', lazy=True, cascade='all,delete')

class Administrado(db.Model):
    id              = db.Column(db.Integer, primary_key=True)
    paciente_id     = db.Column(db.Integer, db.ForeignKey('paciente.id'), nullable=False)
    fecha           = db.Column(db.String(20), nullable=False)
    hora_enfermero  = db.Column(db.String(100))
    nombre_original = db.Column(db.String(300), nullable=False)
    nombre_corregido= db.Column(db.String(300))   # nombre del facturado
    estado          = db.Column(db.String(30), default='pendiente')
    # pendiente / ok / sin_mapeo / revisar / no_facturado
    observacion     = db.Column(db.Text)

class Indicado(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    paciente_id = db.Column(db.Integer, db.ForeignKey('paciente.id'), nullable=False)
    fecha       = db.Column(db.String(20), nullable=False)
    medico      = db.Column(db.String(150))
    droga       = db.Column(db.String(200), nullable=False)
    dosis_via   = db.Column(db.String(300))
    observacion = db.Column(db.Text)

class Facturado(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    paciente_id = db.Column(db.Integer, db.ForeignKey('paciente.id'), nullable=False)
    fecha       = db.Column(db.String(20), nullable=False)
    nombre      = db.Column(db.String(300), nullable=False)
    cantidad    = db.Column(db.Integer, default=1)
    codigo      = db.Column(db.String(30))
    valor       = db.Column(db.String(30))

class Equivalencia(db.Model):
    id              = db.Column(db.Integer, primary_key=True)
    nombre_comercial= db.Column(db.String(300), nullable=False, unique=True)
    nombre_facturado= db.Column(db.String(300), nullable=False)
    activa          = db.Column(db.Boolean, default=True)
    nota            = db.Column(db.Text)
    creado_en       = db.Column(db.DateTime, default=datetime.utcnow)

class LogCambio(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    paciente_id   = db.Column(db.Integer, db.ForeignKey('paciente.id'), nullable=False)
    entidad       = db.Column(db.String(50))   # 'administrado', 'indicado', etc.
    entidad_id    = db.Column(db.Integer)
    campo         = db.Column(db.String(100))
    valor_anterior= db.Column(db.Text)
    valor_nuevo   = db.Column(db.Text)
    fecha         = db.Column(db.DateTime, default=datetime.utcnow)
    usuario       = db.Column(db.String(100), default='sistema')
    descripcion   = db.Column(db.Text)

# ──────────────────────────────────────────
# EQUIVALENCIAS POR DEFECTO
# ──────────────────────────────────────────
EQUIV_DEFAULT = [
    ("HOLOMAGNESIO COMP.REC.X 1 PHOENIX X 2",                   "HOLOMAGNESIO",                                "Citrato de magnesio Phoenix"),
    ("FRIDALIT 100 100 MG F.A.X 1 X 5 ML FADA PHARMA X 1",     "HIDROCORTISONA 100 MG.FCO.AMP.",              "Hidrocortisona Fada Pharma"),
    ("CUBICIN RT 500 MG F.A.X 1 MSD ARGENTINA SR X 2",         "CUBICIN RT 500 MG IV AMPOLLA",                "Daptomicina MSD"),
    ("CUBICIN RT 500 MG F.A.X 1 MSD ARGENTINA SR X 1",         "CUBICIN RT 500 MG IV AMPOLLA",                "Daptomicina MSD"),
    ("CANCIDAS 50 MG VIAL X 1 MSD ARGENTINA SR X 1",           "CANCIDAS 50 MG",                              "Caspofungin MSD - VERIFICAR en facturado"),
    ("AMIODARONA LARJAN 150 MG A.X 1 X 3 ML VEINFAR X 2",      "ATLANSIL COMPRIMIDO",                         "Amiodarona Veinfar IV → Atlansil oral"),
    ("OMEPRAZOL CEVALLOS 20 MG CAPS.X 1 CEVALLOS X 2",         "PROCELAC 20 MG.CAPS.",                        "Omeprazol Cevallos → Procelac"),
    ("MEROPENEM RICHET 500 MG IV INY.F.A.X 1 RICHET X 1",      "MEROEFECTIL 1000 MG F.A.",                    "Meropenem Richet → Meroefectil"),
    ("SOBRIUS 5000 UI/ML F.A.X 1 X5ML FADA PHARMA X 1",        "HEPARINA DUNCAN 5000 U.I.FCO.AMP.X 5 ML",    "Heparina Sobrius Fada → Heparina Duncan"),
    ("ACIDO FOLICO VANNIER 5 MG COMP.X 1 VANNIER X 1",         "ACIFOL 5 MG.COMP.",                           "Acido folico Vannier → Acifol"),
    ("FLOXLEVO 500 MG COMP.REC.X 1 BIOTENK X 1",               "LEVOFLOXACINA 500 MG",                        "Levofloxacina Biotenk 500mg - VERIFICAR en facturado"),
    ("AMINOXIDIN SULBACTAM INY.F.A.X 1 X 20 ML FADA PHARMA X 1","AMPICILINA+SULBACTAM 1,5 G.IM/IV AMP.",     "Aminoxidin Sulbactam Fada → AmpicilinaSulbactam"),
    ("DRENIX 100 100 MG COMP.REC.X 1 OXAPHARMA X 1",           "ACNECLIN 100 MG AP COMP",                     "Minociclina Oxapharma → Acneclin"),
    ("BACTRIM FORTE COMP.X 1 INVESTI X 1",                      "BACTRIM FUERTE COMPRIMIDOS",                  "Bactrim Investi → Bactrim Fuerte"),
    ("CLEXANE 80 MG JGA.PRELL.X 1 SANOFI-AVENTIS X 1",         "CLEXANE 80 MG.AMPOLLA",                       "Enoxaparina Sanofi → Clexane"),
    ("HEMAX 4000UI LIOF.F.A.+J.PRELL BIOSIDUS FARMA X 1",      "HEMAX 2000 U.I.FCO.AMP.",                     "Eritropoyetina Biosidus 4000UI → Hemax 2000UI"),
    ("PROGRAF 1 MG CAPS.X 1 GADOR X 2",                         "PROGRAF 1 MG COMP. RE",                      "Tacrolimus Gador → Prograf"),
    ("UNIFLOX 750 750 MG COMP.REC.X 1 RAFFO X 1",              "LEVOFLOXACINA 750 MG",                        "Levofloxacina Raffo 750mg - VERIFICAR en facturado"),
    ("UNIFLOX 750 MG COMP.REC.X 1 RAFFO X 1",                  "LEVOFLOXACINA 750 MG",                        "Levofloxacina Raffo 750mg"),
    ("LA MEPREDNISONA 4 MG COMP.X 1 BIOTENK X 1",              "DELTISONA B 4 MG",                            "Meprednisona Biotenk → Deltisona B"),
    ("ATLANSIL COMP.X 1 ROEMMERS X 1",                          "ATLANSIL COMPRIMIDO",                         "Amiodarona Roemmers → Atlansil"),
    ("SULFATO DE MAGNESIO 25% A.X 1 X 5 ML FADA PHARMA X 2",   "SULFATO MAGNESIO 25 % X 5 ML",               "Sulfato magnesio Fada"),
]

# ──────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────

def _normalizar_medicamento(txt):
    if not txt:
        return ''
    s = txt.upper().strip()
    s = re.sub(r'\bX\s+', ' ', s)
    s = re.sub(r'[^A-Z0-9\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _buscar_equivalencia(nombre_original, mapa_norm):
    norm_item = _normalizar_medicamento(nombre_original)
    if not norm_item:
        return None

    # 1) Match exacto sobre nombre normalizado
    if norm_item in mapa_norm:
        return mapa_norm[norm_item]

    # 2) Match por inclusión para tolerar OCR que omite separadores
    for k_norm, v_fact in mapa_norm.items():
        if len(k_norm) < 12:
            continue
        if k_norm in norm_item or norm_item in k_norm:
            return v_fact

    # 3) Match por similitud de cadena
    best_ratio = 0.0
    best_value = None
    for k_norm, v_fact in mapa_norm.items():
        ratio = SequenceMatcher(None, norm_item, k_norm).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_value = v_fact
    if best_ratio >= 0.78:
        return best_value
    return None


def _coincide_nombre_facturado(nombre_objetivo, nombre_candidato):
    """Match tolerante para comparar nombre corregido vs nombre facturado del PDF."""
    a = _normalizar_medicamento(nombre_objetivo)
    b = _normalizar_medicamento(nombre_candidato)
    if not a or not b:
        return False
    if a == b:
        return True

    if len(a) >= 10 and len(b) >= 10 and (a in b or b in a):
        return True

    ta = set(a.split())
    tb = set(b.split())
    if ta and tb:
        inter = ta & tb
        cobertura = len(inter) / max(1, min(len(ta), len(tb)))
        if len(inter) >= 2 and cobertura >= 0.5:
            return True

    return SequenceMatcher(None, a, b).ratio() >= 0.72


def _extraer_mg(nombre):
    """Extrae la primera potencia en mg del texto, si existe."""
    txt = _normalizar_medicamento(nombre)
    if not txt:
        return None
    m = re.search(r'\b(\d+(?:[\.,]\d+)?)\s*MG\b', txt)
    if not m:
        return None
    try:
        return float(m.group(1).replace(',', '.'))
    except Exception:
        return None


def _sin_potencia(nombre):
    txt = _normalizar_medicamento(nombre)
    txt = re.sub(r'\b\d+(?:[\.,]\d+)?\s*MG\b', ' ', txt)
    txt = re.sub(r'\b\d+(?:[\.,]\d+)?\b', ' ', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt


def _es_ruido_facturado(nombre):
    """Filtra descartables/ruido típico del OCR en facturados."""
    n = _normalizar_medicamento(nombre)
    if not n:
        return True
    ruido_tokens = [
        'DESCART', 'TRANSP', 'SOL FIS', 'RINGER', 'SONDA', 'CATETER',
        'GASA', 'APOSITO', 'GUANTE', 'AGUJA', 'JERINGA'
    ]
    return any(tok in n for tok in ruido_tokens)


def _coincide_facturado_con_cantidad(nombre_objetivo, nombre_facturado, cantidad=1):
    """Match por nombre y, si aplica, por potencia total (ej: 100mg vs 2x50mg)."""
    if _coincide_nombre_facturado(nombre_objetivo, nombre_facturado):
        return True

    base_obj = _sin_potencia(nombre_objetivo)
    base_fac = _sin_potencia(nombre_facturado)
    if not base_obj or not base_fac:
        return False

    if not _coincide_nombre_facturado(base_obj, base_fac):
        return False

    mg_obj = _extraer_mg(nombre_objetivo)
    mg_fac = _extraer_mg(nombre_facturado)
    if mg_obj is None or mg_fac is None:
        return False

    try:
        cant = max(1, int(cantidad or 1))
    except Exception:
        cant = 1

    total_fac = mg_fac * cant
    if total_fac <= 0:
        return False

    # Tolerancia para redondeos/variantes OCR
    return abs(mg_obj - total_fac) / max(mg_obj, total_fac) <= 0.20


def aplicar_equivalencias(paciente_id):
    """Aplica la tabla de equivalencias a todos los administrados de un paciente"""
    equivs = Equivalencia.query.filter_by(activa=True).all()
    mapa_norm = {
        _normalizar_medicamento(e.nombre_comercial): e.nombre_facturado
        for e in equivs
        if _normalizar_medicamento(e.nombre_comercial)
    }
    items = Administrado.query.filter_by(paciente_id=paciente_id).all()
    cambios = 0
    for item in items:
        nuevo = _buscar_equivalencia(item.nombre_original, mapa_norm)
        if nuevo and item.nombre_corregido != nuevo:
            viejo = item.nombre_corregido
            item.nombre_corregido = nuevo
            item.estado = 'ok'
            log = LogCambio(
                paciente_id=paciente_id, entidad='administrado',
                entidad_id=item.id, campo='nombre_corregido',
                valor_anterior=viejo or '', valor_nuevo=nuevo,
                usuario='sistema',
                descripcion=f'Mapeo automático: {item.nombre_original[:60]} → {nuevo}'
            )
            db.session.add(log)
            cambios += 1
        elif not nuevo and item.estado == 'pendiente':
            item.estado = 'sin_mapeo'
    db.session.commit()
    return cambios

def stats_paciente(pac_id):
    adm   = Administrado.query.filter_by(paciente_id=pac_id).all()
    ind   = Indicado.query.filter_by(paciente_id=pac_id).all()
    fact  = Facturado.query.filter_by(paciente_id=pac_id).all()
    return {
        'total_adm':      len(adm),
        'ok':             sum(1 for a in adm if a.estado == 'ok'),
        'sin_mapeo':      sum(1 for a in adm if a.estado == 'sin_mapeo'),
        'revisar':        sum(1 for a in adm if a.estado == 'revisar'),
        'no_facturado':   sum(1 for a in adm if a.estado == 'no_facturado'),
        'pendiente':      sum(1 for a in adm if a.estado == 'pendiente'),
        'total_ind':      len(ind),
        'total_fact':     len(fact),
    }

# ──────────────────────────────────────────
# RUTAS PRINCIPALES
# ──────────────────────────────────────────

@app.route('/')
def index():
    pacientes = Paciente.query.order_by(Paciente.creado_en.desc()).all()
    stats = {p.id: stats_paciente(p.id) for p in pacientes}
    total_equiv = Equivalencia.query.filter_by(activa=True).count()
    return render_template('index.html', pacientes=pacientes, stats=stats, total_equiv=total_equiv)

@app.route('/paciente/nuevo', methods=['GET','POST'])
def nuevo_paciente():
    if request.method == 'POST':
        pac = Paciente(
            nombre=request.form['nombre'].upper(),
            dni=request.form.get('dni',''),
            fecha_ingreso=request.form.get('fecha_ingreso',''),
            fecha_egreso=request.form.get('fecha_egreso',''),
            hc='',
            plan='',
            cama='',
            notas='',
        )
        db.session.add(pac)
        db.session.commit()
        flash(f'Paciente {pac.nombre} creado correctamente.', 'success')
        return redirect(url_for('paciente_detalle', pid=pac.id))
    return render_template('nuevo_paciente.html')

@app.route('/paciente/<int:pid>')
def paciente_detalle(pid):
    pac = Paciente.query.get_or_404(pid)
    st  = stats_paciente(pid)
    return render_template('paciente_detalle.html', pac=pac, st=st)

@app.route('/paciente/<int:pid>/editar', methods=['POST'])
def editar_paciente(pid):
    pac = Paciente.query.get_or_404(pid)

    nombre_nuevo = request.form.get('nombre', '').strip().upper()
    if not nombre_nuevo:
        flash('El nombre del paciente es obligatorio.', 'danger')
        return redirect(url_for('paciente_detalle', pid=pid))

    campos = {
        'nombre': nombre_nuevo,
        'dni': request.form.get('dni', '').strip(),
        'fecha_ingreso': request.form.get('fecha_ingreso', '').strip(),
        'fecha_egreso': request.form.get('fecha_egreso', '').strip(),
        'hc': '',
        'plan': '',
        'cama': '',
        'notas': '',
    }

    cambios = 0
    for campo, nuevo_valor in campos.items():
        valor_actual = getattr(pac, campo) or ''
        if str(valor_actual) != str(nuevo_valor):
            db.session.add(LogCambio(
                paciente_id=pid,
                entidad='paciente',
                entidad_id=pid,
                campo=campo,
                valor_anterior=str(valor_actual),
                valor_nuevo=str(nuevo_valor),
                usuario='usuario',
                descripcion=f'Edicion de paciente: {campo}'
            ))
            setattr(pac, campo, nuevo_valor)
            cambios += 1

    if cambios:
        db.session.commit()
        flash('Datos del paciente actualizados correctamente.', 'success')
    else:
        flash('No hubo cambios para guardar.', 'danger')

    return redirect(url_for('paciente_detalle', pid=pid))

@app.route('/paciente/<int:pid>/eliminar', methods=['POST'])
def eliminar_paciente(pid):
    pac = Paciente.query.get_or_404(pid)
    nombre = pac.nombre
    db.session.delete(pac)
    db.session.commit()
    flash(f'Paciente {nombre} eliminado correctamente.', 'success')
    return redirect(url_for('index'))

@app.route('/paciente/<int:pid>/stats_json')
def stats_json(pid):
    return jsonify(stats_paciente(pid))

# ──────── ADMINISTRADOS ────────

@app.route('/paciente/<int:pid>/administrados')
def administrados(pid):
    pac  = Paciente.query.get_or_404(pid)
    adms = Administrado.query.filter_by(paciente_id=pid).order_by(
           Administrado.fecha, Administrado.hora_enfermero).all()
    equivs = [e.nombre_facturado for e in Equivalencia.query.filter_by(activa=True).all()]
    return render_template('administrados.html', pac=pac, adms=adms, equivs=sorted(set(equivs)))

@app.route('/paciente/<int:pid>/administrados/nuevo', methods=['POST'])
def nuevo_administrado(pid):
    data = request.get_json()
    adm = Administrado(
        paciente_id=pid,
        fecha=data['fecha'],
        hora_enfermero=data.get('hora_enfermero',''),
        nombre_original=data['nombre_original'].strip(),
        nombre_corregido=data.get('nombre_corregido','').strip() or None,
        estado=data.get('estado','pendiente'),
        observacion=data.get('observacion',''),
    )
    db.session.add(adm)
    db.session.flush()
    log = LogCambio(paciente_id=pid, entidad='administrado', entidad_id=adm.id,
                    campo='creacion', valor_anterior='', valor_nuevo=adm.nombre_original,
                    usuario=data.get('usuario','usuario'),
                    descripcion=f'Administrado agregado: {adm.nombre_original[:80]}')
    db.session.add(log)
    db.session.commit()
    return jsonify({'ok': True, 'id': adm.id})

@app.route('/paciente/<int:pid>/administrados/<int:aid>', methods=['PUT'])
def editar_administrado(pid, aid):
    adm  = Administrado.query.get_or_404(aid)
    data = request.get_json()
    logs = []
    for campo in ['nombre_corregido','estado','observacion','fecha','hora_enfermero','nombre_original']:
        if campo in data:
            viejo = getattr(adm, campo) or ''
            nuevo = data[campo]
            if str(viejo) != str(nuevo):
                logs.append(LogCambio(
                    paciente_id=pid, entidad='administrado', entidad_id=aid,
                    campo=campo, valor_anterior=str(viejo), valor_nuevo=str(nuevo),
                    usuario=data.get('usuario','usuario'),
                    descripcion=f'Edición manual: {campo} → {str(nuevo)[:80]}'
                ))
                setattr(adm, campo, nuevo)
    for l in logs: db.session.add(l)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/paciente/<int:pid>/administrados/<int:aid>', methods=['DELETE'])
def eliminar_administrado(pid, aid):
    adm = Administrado.query.get_or_404(aid)
    log = LogCambio(paciente_id=pid, entidad='administrado', entidad_id=aid,
                    campo='eliminacion', valor_anterior=adm.nombre_original, valor_nuevo='',
                    usuario='usuario', descripcion=f'Eliminado: {adm.nombre_original[:80]}')
    db.session.add(log)
    db.session.delete(adm)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/paciente/<int:pid>/aplicar_equiv', methods=['POST'])
def aplicar_equiv(pid):
    n = aplicar_equivalencias(pid)
    return jsonify({'ok': True, 'cambios': n})

# ──────── INDICADOS ────────

@app.route('/paciente/<int:pid>/indicados')
def indicados(pid):
    pac  = Paciente.query.get_or_404(pid)
    inds = Indicado.query.filter_by(paciente_id=pid).order_by(
           Indicado.fecha, Indicado.medico).all()
    return render_template('indicados.html', pac=pac, inds=inds)

@app.route('/paciente/<int:pid>/indicados/nuevo', methods=['POST'])
def nuevo_indicado(pid):
    data = request.get_json()
    ind = Indicado(
        paciente_id=pid,
        fecha=data['fecha'],
        medico=data.get('medico',''),
        droga=data['droga'].strip(),
        dosis_via=data.get('dosis_via',''),
        observacion=data.get('observacion',''),
    )
    db.session.add(ind)
    db.session.flush()
    log = LogCambio(paciente_id=pid, entidad='indicado', entidad_id=ind.id,
                    campo='creacion', valor_anterior='', valor_nuevo=ind.droga,
                    usuario=data.get('usuario','usuario'),
                    descripcion=f'Indicado agregado: {ind.droga[:80]}')
    db.session.add(log)
    db.session.commit()
    return jsonify({'ok': True, 'id': ind.id})

@app.route('/paciente/<int:pid>/indicados/<int:iid>', methods=['PUT'])
def editar_indicado(pid, iid):
    ind  = Indicado.query.get_or_404(iid)
    data = request.get_json()
    for campo in ['fecha','medico','droga','dosis_via','observacion']:
        if campo in data:
            viejo = getattr(ind, campo) or ''
            nuevo = data[campo]
            if str(viejo) != str(nuevo):
                log = LogCambio(paciente_id=pid, entidad='indicado', entidad_id=iid,
                                campo=campo, valor_anterior=str(viejo), valor_nuevo=str(nuevo),
                                usuario=data.get('usuario','usuario'),
                                descripcion=f'Edición: {campo} → {str(nuevo)[:60]}')
                db.session.add(log)
                setattr(ind, campo, nuevo)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/paciente/<int:pid>/indicados/<int:iid>', methods=['DELETE'])
def eliminar_indicado(pid, iid):
    ind = Indicado.query.get_or_404(iid)
    log = LogCambio(paciente_id=pid, entidad='indicado', entidad_id=iid,
                    campo='eliminacion', valor_anterior=ind.droga, valor_nuevo='',
                    usuario='usuario', descripcion=f'Eliminado indicado: {ind.droga[:80]}')
    db.session.add(log)
    db.session.delete(ind)
    db.session.commit()
    return jsonify({'ok': True})

# ──────── FACTURADOS ────────

@app.route('/paciente/<int:pid>/facturados')
def facturados(pid):
    pac   = Paciente.query.get_or_404(pid)
    facts = Facturado.query.filter_by(paciente_id=pid).order_by(Facturado.fecha).all()
    return render_template('facturados.html', pac=pac, facts=facts)

@app.route('/paciente/<int:pid>/facturados/nuevo', methods=['POST'])
def nuevo_facturado(pid):
    data = request.get_json()
    fact = Facturado(
        paciente_id=pid,
        fecha=data['fecha'],
        nombre=data['nombre'].strip(),
        cantidad=int(data.get('cantidad', 1)),
        codigo=data.get('codigo',''),
        valor=data.get('valor',''),
    )
    db.session.add(fact)
    db.session.flush()
    log = LogCambio(paciente_id=pid, entidad='facturado', entidad_id=fact.id,
                    campo='creacion', valor_anterior='', valor_nuevo=fact.nombre,
                    usuario=data.get('usuario','usuario'),
                    descripcion=f'Facturado agregado: {fact.nombre[:80]}')
    db.session.add(log)
    db.session.commit()
    return jsonify({'ok': True, 'id': fact.id})

@app.route('/paciente/<int:pid>/facturados/<int:fid>', methods=['PUT'])
def editar_facturado(pid, fid):
    fact = Facturado.query.get_or_404(fid)
    data = request.get_json()
    for campo in ['fecha','nombre','cantidad','codigo','valor']:
        if campo in data:
            viejo = getattr(fact, campo)
            nuevo = data[campo]
            if str(viejo) != str(nuevo):
                log = LogCambio(paciente_id=pid, entidad='facturado', entidad_id=fid,
                                campo=campo, valor_anterior=str(viejo), valor_nuevo=str(nuevo),
                                usuario=data.get('usuario','usuario'),
                                descripcion=f'Edición facturado: {campo}')
                db.session.add(log)
                setattr(fact, campo, nuevo)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/paciente/<int:pid>/facturados/<int:fid>', methods=['DELETE'])
def eliminar_facturado(pid, fid):
    fact = Facturado.query.get_or_404(fid)
    log = LogCambio(paciente_id=pid, entidad='facturado', entidad_id=fid,
                    campo='eliminacion', valor_anterior=fact.nombre, valor_nuevo='',
                    usuario='usuario', descripcion=f'Eliminado facturado: {fact.nombre[:80]}')
    db.session.add(log)
    db.session.delete(fact)
    db.session.commit()
    return jsonify({'ok': True})

# ──────── CONTROL CRUZADO ────────

@app.route('/paciente/<int:pid>/control')
def control(pid):
    pac   = Paciente.query.get_or_404(pid)
    adms  = Administrado.query.filter_by(paciente_id=pid).order_by(Administrado.fecha).all()
    facts_all = Facturado.query.filter_by(paciente_id=pid).order_by(Facturado.fecha).all()
    facts = [f for f in facts_all if not _es_ruido_facturado(f.nombre)]
    inds  = Indicado.query.filter_by(paciente_id=pid).order_by(Indicado.fecha).all()

    # Facturados agrupados por fecha (solo medicamentos, sin descartables/ruido)
    fact_por_fecha = {}
    for f in facts:
        key = f.fecha
        if key not in fact_por_fecha:
            fact_por_fecha[key] = []
        fact_por_fecha[key].append(f)
    fact_todos = facts

    # Construir filas de control
    rows = []
    for adm in adms:
        nombre_fact = adm.nombre_corregido or ''
        fecha = adm.fecha
        en_facturado_fecha = any(
            nombre_fact and _coincide_facturado_con_cantidad(nombre_fact, f.nombre, f.cantidad)
            for f in fact_por_fecha.get(fecha, [])
        ) if nombre_fact else False
        en_facturado_global = any(
            nombre_fact and _coincide_facturado_con_cantidad(nombre_fact, f.nombre, f.cantidad)
            for f in fact_todos
        ) if nombre_fact else False
        en_facturado = en_facturado_fecha or en_facturado_global

        if adm.estado == 'ok' and en_facturado:
            estado_ctrl = 'completo'
            badge = 'success'
            label = 'OK Completo' if en_facturado_fecha else 'OK (fecha distinta)'
        elif adm.estado == 'ok' and not en_facturado:
            estado_ctrl = 'adm_no_fact'
            badge = 'warning'
            label = 'Adm. no facturado'
        elif adm.estado == 'sin_mapeo':
            estado_ctrl = 'sin_mapeo'
            badge = 'danger'
            label = 'Sin equivalencia'
        elif adm.estado == 'revisar':
            estado_ctrl = 'revisar'
            badge = 'secondary'
            label = 'Revisar'
        else:
            estado_ctrl = 'pendiente'
            badge = 'info'
            label = 'Pendiente'

        rows.append({
            'id': adm.id,
            'fecha': fecha,
            'hora_enfermero': adm.hora_enfermero,
            'nombre_original': adm.nombre_original,
            'nombre_corregido': nombre_fact,
            'observacion': adm.observacion or '',
            'estado_ctrl': estado_ctrl,
            'badge': badge,
            'label': label,
            'en_facturado': en_facturado,
        })

    # Facturados sin administrado correspondiente
    adm_corregidos = set(a.nombre_corregido for a in adms if a.nombre_corregido)
    fact_sin_adm = []
    for f in facts:
        if not any(_coincide_facturado_con_cantidad(ac, f.nombre, f.cantidad) for ac in adm_corregidos):
            fact_sin_adm.append(f)

    st = stats_paciente(pid)
    return render_template('control.html', pac=pac, rows=rows, inds=inds,
                           fact_sin_adm=fact_sin_adm, st=st)

# ──────── LOG DE CAMBIOS ────────

@app.route('/paciente/<int:pid>/log')
def log_cambios(pid):
    pac  = Paciente.query.get_or_404(pid)
    logs = LogCambio.query.filter_by(paciente_id=pid)\
                    .order_by(LogCambio.fecha.desc()).all()
    return render_template('log_cambios.html', pac=pac, logs=logs)

# ──────── EQUIVALENCIAS ────────

@app.route('/equivalencias')
def equivalencias():
    equivs = Equivalencia.query.order_by(Equivalencia.nombre_comercial).all()
    return render_template('equivalencias.html', equivs=equivs)

@app.route('/equivalencias/nueva', methods=['POST'])
def nueva_equivalencia():
    data = request.get_json()
    exist = Equivalencia.query.filter_by(nombre_comercial=data['nombre_comercial'].strip()).first()
    if exist:
        exist.nombre_facturado = data['nombre_facturado'].strip()
        exist.nota = data.get('nota','')
        exist.activa = True
    else:
        eq = Equivalencia(
            nombre_comercial=data['nombre_comercial'].strip(),
            nombre_facturado=data['nombre_facturado'].strip(),
            nota=data.get('nota',''),
        )
        db.session.add(eq)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/equivalencias/<int:eid>', methods=['PUT'])
def editar_equivalencia(eid):
    eq = Equivalencia.query.get_or_404(eid)
    data = request.get_json()
    for campo in ['nombre_comercial','nombre_facturado','nota','activa']:
        if campo in data:
            setattr(eq, campo, data[campo])
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/equivalencias/<int:eid>', methods=['DELETE'])
def eliminar_equivalencia(eid):
    eq = Equivalencia.query.get_or_404(eid)
    db.session.delete(eq)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/equivalencias/datos')
def datos_equivalencias():
    equivs = Equivalencia.query.filter_by(activa=True).order_by(Equivalencia.nombre_comercial).all()
    return jsonify([{
        'id': e.id, 'nombre_comercial': e.nombre_comercial,
        'nombre_facturado': e.nombre_facturado, 'nota': e.nota or '', 'activa': e.activa
    } for e in equivs])

# ──────── IMPORTAR DESDE EXCEL ────────

def _detectar_col(headers, candidatos):
    """Devuelve el índice de la primera columna cuyo nombre coincida con algún candidato."""
    for i, h in enumerate(headers):
        h_norm = str(h).lower().strip()
        for c in candidatos:
            if c in h_norm:
                return i
    return None

def _val(row, idx, default=''):
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    return str(v).strip() if v is not None else default

def _parse_excel(filepath, tipo):
    """
    Lee un Excel y devuelve lista de dicts según el tipo:
    tipo = 'administrados' | 'facturados' | 'indicados'
    Detecta cabeceras automáticamente; omite filas vacías.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Buscar fila de cabecera (primera fila no vacía)
    headers = []
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 2:
            headers = [str(c).lower().strip() if c else '' for c in row]
            header_row_idx = i + 1
            break

    if not headers:
        return [], "No se encontró fila de cabecera con al menos 2 columnas."

    results = []

    if tipo == 'administrados':
        ci_fecha    = _detectar_col(headers, ['fecha', 'date', 'dia', 'día'])
        ci_enfermero= _detectar_col(headers, ['enfermero', 'hora', 'profesional', 'turno', 'nurse'])
        ci_nombre   = _detectar_col(headers, ['medicamento', 'droga', 'nombre', 'med', 'farmaco', 'fármaco',
                                               'producto', 'descripcion', 'descripción', 'drug'])
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
            fecha  = _val(row, ci_fecha)
            nombre = _val(row, ci_nombre)
            if not fecha or not nombre or nombre.lower() in ('none','nan',''):
                continue
            results.append({
                'fecha':          fecha,
                'hora_enfermero': _val(row, ci_enfermero),
                'nombre_original': nombre,
            })

    elif tipo == 'facturados':
        ci_fecha    = _detectar_col(headers, ['fecha', 'date', 'dia', 'día'])
        ci_nombre   = _detectar_col(headers, ['nombre', 'descripcion', 'descripción', 'medicamento',
                                               'producto', 'item', 'droga'])
        ci_cantidad = _detectar_col(headers, ['cantidad', 'cant', 'qty', 'unidades', 'q'])
        ci_codigo   = _detectar_col(headers, ['codigo', 'código', 'cod', 'code', 'art'])
        ci_valor    = _detectar_col(headers, ['valor', 'precio', 'importe', 'monto', 'price'])
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
            fecha  = _val(row, ci_fecha)
            nombre = _val(row, ci_nombre)
            if not fecha or not nombre or nombre.lower() in ('none','nan',''):
                continue
            cant = _val(row, ci_cantidad, '1')
            try:
                cant = int(float(cant))
            except:
                cant = 1
            results.append({
                'fecha':    fecha,
                'nombre':   nombre,
                'cantidad': cant,
                'codigo':   _val(row, ci_codigo),
                'valor':    _val(row, ci_valor),
            })

    elif tipo == 'indicados':
        ci_fecha  = _detectar_col(headers, ['fecha', 'date', 'dia', 'día'])
        ci_medico = _detectar_col(headers, ['medico', 'médico', 'doctor', 'dr', 'profesional'])
        ci_droga  = _detectar_col(headers, ['droga', 'medicamento', 'nombre', 'indicacion',
                                             'indicación', 'farmaco', 'fármaco', 'drug'])
        ci_dosis  = _detectar_col(headers, ['dosis', 'via', 'vía', 'dose', 'posologia', 'posología'])
        ci_obs    = _detectar_col(headers, ['observacion', 'observación', 'obs', 'nota', 'note'])
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
            fecha = _val(row, ci_fecha)
            droga = _val(row, ci_droga)
            if not fecha or not droga or droga.lower() in ('none','nan',''):
                continue
            results.append({
                'fecha':    fecha,
                'medico':   _val(row, ci_medico),
                'droga':    droga,
                'dosis_via':_val(row, ci_dosis),
                'observacion': _val(row, ci_obs),
            })

    return results, None


@app.route('/paciente/<int:pid>/cargar', methods=['GET'])
def cargar_datos(pid):
    pac = Paciente.query.get_or_404(pid)
    return render_template('cargar.html', pac=pac)


@app.route('/paciente/<int:pid>/cargar_excel', methods=['POST'])
def cargar_excel(pid):
    pac  = Paciente.query.get_or_404(pid)
    tipo = request.form.get('tipo')  # administrados | facturados | indicados
    archivo = request.files.get('archivo')

    if not archivo or not tipo:
        return jsonify({'ok': False, 'error': 'Falta archivo o tipo'})

    ext = os.path.splitext(archivo.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'ok': False, 'error': 'Solo se aceptan archivos .xlsx o .xls'})

    tmp = os.path.join(app.config['UPLOAD_FOLDER'], f'tmp_{pid}_{tipo}{ext}')
    archivo.save(tmp)

    try:
        rows, err = _parse_excel(tmp, tipo)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error al leer el archivo: {e}'})
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)

    if err:
        return jsonify({'ok': False, 'error': err})
    if not rows:
        return jsonify({'ok': False, 'error': 'No se encontraron filas con datos válidos. Revisá que el archivo tenga las columnas correctas.'})

    # Insertar en base de datos
    n = 0
    if tipo == 'administrados':
        for r in rows:
            db.session.add(Administrado(
                paciente_id=pid, fecha=r['fecha'],
                hora_enfermero=r['hora_enfermero'],
                nombre_original=r['nombre_original'],
                estado='pendiente'
            ))
            n += 1
        db.session.commit()
        aplicar_equivalencias(pid)

    elif tipo == 'facturados':
        for r in rows:
            db.session.add(Facturado(
                paciente_id=pid, fecha=r['fecha'], nombre=r['nombre'],
                cantidad=r['cantidad'], codigo=r['codigo'], valor=r['valor']
            ))
            n += 1
        db.session.commit()

    elif tipo == 'indicados':
        for r in rows:
            db.session.add(Indicado(
                paciente_id=pid, fecha=r['fecha'], medico=r['medico'],
                droga=r['droga'], dosis_via=r['dosis_via'],
                observacion=r.get('observacion','')
            ))
            n += 1
        db.session.commit()

    log = LogCambio(paciente_id=pid, entidad=tipo, entidad_id=0,
                    campo='importacion_excel', valor_anterior='', valor_nuevo=str(n),
                    usuario='usuario',
                    descripcion=f'Importación Excel {tipo}: {n} registros')
    db.session.add(log)
    db.session.commit()

    return jsonify({'ok': True, 'importados': n, 'tipo': tipo})


@app.route('/paciente/<int:pid>/cargar_pdfs', methods=['POST'])
def cargar_pdfs(pid):
    Paciente.query.get_or_404(pid)
    pdf_ind = request.files.get('indicaciones_pdf')
    pdf_fact = request.files.get('facturados_pdf')
    reemplazar = request.form.get('reemplazar', '1') in ('1', 'true', 'True', 'on')

    if not pdf_ind or not pdf_fact:
        return jsonify({'ok': False, 'error': 'Debes subir ambos archivos PDF.'})

    ext_ind = os.path.splitext(pdf_ind.filename or '')[1].lower()
    ext_fact = os.path.splitext(pdf_fact.filename or '')[1].lower()
    if ext_ind != '.pdf' or ext_fact != '.pdf':
        return jsonify({'ok': False, 'error': 'Los dos archivos deben ser PDF.'})

    job_id = uuid.uuid4().hex
    tmp_ind = os.path.join(app.config['UPLOAD_FOLDER'], f'tmp_{pid}_{job_id}_indicaciones.pdf')
    tmp_fact = os.path.join(app.config['UPLOAD_FOLDER'], f'tmp_{pid}_{job_id}_facturados.pdf')

    pdf_ind.save(tmp_ind)
    pdf_fact.save(tmp_fact)

    # Guardar una copia estable del PDF de indicaciones para exportaciones posteriores.
    try:
        ruta_estable_ind = os.path.join(app.config['UPLOAD_FOLDER'], f'paciente_{pid}_indicaciones.pdf')
        shutil.copyfile(tmp_ind, ruta_estable_ind)
    except Exception:
        pass

    with PDF_JOBS_LOCK:
        PDF_JOBS[job_id] = {
            'status': 'queued',
            'progress': 0,
            'message': 'En cola...'
        }

    thread = threading.Thread(
        target=_procesar_pdfs_job,
        args=(pid, job_id, tmp_ind, tmp_fact, reemplazar),
        daemon=True
    )
    thread.start()

    return jsonify({'ok': True, 'job_id': job_id})


@app.route('/paciente/<int:pid>/cargar_pdfs_estado/<job_id>', methods=['GET'])
def cargar_pdfs_estado(pid, job_id):
    Paciente.query.get_or_404(pid)
    with PDF_JOBS_LOCK:
        job = PDF_JOBS.get(job_id)
    if not job:
        return jsonify({'ok': False, 'error': 'Job no encontrado'})
    return jsonify({'ok': True, **job})


@app.route('/paciente/<int:pid>/borrar_datos', methods=['POST'])
def borrar_datos(pid):
    """Borra todos los registros de un tipo para re-importar limpio."""
    tipo = request.get_json().get('tipo')
    pac = Paciente.query.get_or_404(pid)
    if tipo == 'administrados':
        Administrado.query.filter_by(paciente_id=pid).delete()
    elif tipo == 'facturados':
        Facturado.query.filter_by(paciente_id=pid).delete()
    elif tipo == 'indicados':
        Indicado.query.filter_by(paciente_id=pid).delete()
    else:
        return jsonify({'ok': False, 'error': 'Tipo inválido'})
    db.session.commit()
    return jsonify({'ok': True})


# ──────── EXPORTAR EXCEL ────────

@app.route('/paciente/<int:pid>/exportar_excel')
def exportar_excel(pid):
    from utils.excel_gen import generar_excel
    pac   = Paciente.query.get_or_404(pid)
    adms  = Administrado.query.filter_by(paciente_id=pid).order_by(Administrado.fecha).all()
    inds  = Indicado.query.filter_by(paciente_id=pid).order_by(Indicado.fecha).all()
    facts = Facturado.query.filter_by(paciente_id=pid).order_by(Facturado.fecha).all()
    logs  = LogCambio.query.filter_by(paciente_id=pid).order_by(LogCambio.fecha.desc()).all()
    buf = generar_excel(pac, adms, inds, facts, logs)
    nombre = f"Control_{pac.nombre.replace(' ','_')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/paciente/<int:pid>/exportar_pdf_comparacion')
def exportar_pdf_comparacion(pid):
    import fitz

    pac = Paciente.query.get_or_404(pid)
    adms = Administrado.query.filter_by(paciente_id=pid).order_by(Administrado.fecha, Administrado.id).all()
    facts_all = Facturado.query.filter_by(paciente_id=pid).order_by(Facturado.fecha, Facturado.id).all()
    facts = [f for f in facts_all if not _es_ruido_facturado(f.nombre)]

    doc = fitz.open()

    # Base: PDF original de indicaciones del paciente
    ruta_base = _buscar_pdf_indicaciones_paciente(pid)
    if ruta_base and os.path.exists(ruta_base):
        try:
            src = fitz.open(ruta_base)
            doc.insert_pdf(src)
            src.close()
        except Exception:
            pass

    # Agregar al final solo la tabla de equivalencias
    page = doc.new_page()
    y = 40
    left = 40
    max_y = page.rect.height - 40

    def nueva_pagina_tabla():
        nonlocal page, y
        page = doc.new_page()
        y = 40
        page.insert_text((left, y), 'TABLA DE EQUIVALENCIAS - MEDICAMENTOS ADMINISTRADOS', fontsize=12, fontname='helv')
        y += 18
        page.insert_text((left, y), f'{pac.nombre}', fontsize=9, fontname='helv')
        y += 14
        page.insert_text((left, y), 'FECHA', fontsize=9, fontname='helv')
        page.insert_text((left + 55, y), 'NOMBRE COMERCIAL (ADMINISTRADO)', fontsize=9, fontname='helv')
        page.insert_text((left + 315, y), 'NOMBRE EN LIQUIDACION (FACTURADO)', fontsize=9, fontname='helv')
        y += 14

    nueva_pagina_tabla()

    filas = []
    vistos = set()
    for a in adms:
        if not a.nombre_original or not a.nombre_corregido:
            continue

        # Mostrar solo equivalencias con correspondencia en facturados.
        if not any(_coincide_facturado_con_cantidad(a.nombre_corregido, f.nombre, f.cantidad) for f in facts):
            continue

        key = (a.fecha or '', a.nombre_original, a.nombre_corregido)
        if key in vistos:
            continue
        vistos.add(key)
        filas.append(key)

    def _fecha_key(s):
        try:
            d, m, y2 = (s or '').split('/')
            if len(y2) == 2:
                y2 = f'20{y2}'
            return (int(y2), int(m), int(d))
        except Exception:
            return (9999, 12, 31)

    filas = sorted(filas, key=lambda x: (_fecha_key(x[0]), _normalizar_medicamento(x[1])))

    for fecha, orig, corr in filas:
        if y > max_y:
            nueva_pagina_tabla()

        f_lines = textwrap.wrap(fecha or '-', width=10) or ['']
        o_lines = textwrap.wrap(orig, width=46) or ['']
        c_lines = textwrap.wrap(corr, width=34) or ['']
        lines = max(len(f_lines), len(o_lines), len(c_lines))
        for i in range(lines):
            if y > max_y:
                nueva_pagina_tabla()
            page.insert_text((left, y), f_lines[i] if i < len(f_lines) else '', fontsize=8, fontname='helv')
            page.insert_text((left + 55, y), o_lines[i] if i < len(o_lines) else '', fontsize=8, fontname='helv')
            page.insert_text((left + 315, y), c_lines[i] if i < len(c_lines) else '', fontsize=8, fontname='helv')
            y += 10
        y += 2

    pdf_bytes = doc.tobytes()
    doc.close()

    buf = io.BytesIO(pdf_bytes)
    buf.seek(0)
    nombre = f'Indicaciones_con_equivalencias_{pac.nombre.replace(" ", "_")}.pdf'
    return send_file(buf, as_attachment=True, download_name=nombre, mimetype='application/pdf')

# ──────── IMPORT MASIVO (JSON) ────────

@app.route('/paciente/<int:pid>/importar', methods=['POST'])
def importar_datos(pid):
    """Importa datos en JSON: {administrados:[...], indicados:[...], facturados:[...]}"""
    data = request.get_json()
    pac = Paciente.query.get_or_404(pid)
    n = {'adm': 0, 'ind': 0, 'fact': 0}

    for a in data.get('administrados', []):
        adm = Administrado(paciente_id=pid, fecha=a['fecha'],
                           hora_enfermero=a.get('hora_enfermero',''),
                           nombre_original=a['nombre_original'],
                           nombre_corregido=a.get('nombre_corregido'),
                           estado='pendiente')
        db.session.add(adm)
        n['adm'] += 1

    for i in data.get('indicados', []):
        ind = Indicado(paciente_id=pid, fecha=i['fecha'],
                       medico=i.get('medico',''), droga=i['droga'],
                       dosis_via=i.get('dosis_via',''))
        db.session.add(ind)
        n['ind'] += 1

    for f in data.get('facturados', []):
        fact = Facturado(paciente_id=pid, fecha=f['fecha'],
                         nombre=f['nombre'], cantidad=int(f.get('cantidad',1)),
                         codigo=f.get('codigo',''), valor=f.get('valor',''))
        db.session.add(fact)
        n['fact'] += 1

    db.session.commit()
    aplicar_equivalencias(pid)

    log = LogCambio(paciente_id=pid, entidad='importacion', entidad_id=0,
                    campo='importacion_masiva', valor_anterior='', valor_nuevo=str(n),
                    usuario='sistema', descripcion=f'Importacion: {n}')
    db.session.add(log)
    db.session.commit()
    return jsonify({'ok': True, 'importados': n})

# ──────────────────────────────────────────
# INIT
# ──────────────────────────────────────────

def init_db():
    db.create_all()
    if Equivalencia.query.count() == 0:
        for com, fact, nota in EQUIV_DEFAULT:
            db.session.add(Equivalencia(nombre_comercial=com, nombre_facturado=fact, nota=nota))
        db.session.commit()
        print(f"[init] {len(EQUIV_DEFAULT)} equivalencias cargadas.")

if __name__ == '__main__':
    with app.app_context():
        init_db()
    print("\n" + "="*50)
    print("  MedControl arrancando...")
    print("  Abrir: http://localhost:5000")
    print("="*50 + "\n")
    app.run(debug=True, port=5000)
