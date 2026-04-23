import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ESTADO_FILL = {
    'ok':           'C6EFCE',
    'sin_mapeo':    'FFC7CE',
    'revisar':      'FFEB9C',
    'no_facturado': 'FFCC99',
    'pendiente':    'DDEBF7',
}
BORDE = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def cel(ws, row, col, val, fill=None, bold=False, wrap=True, align='left', fontsize=10):
    c = ws.cell(row=row, column=col, value=val)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if bold:
        c.font = Font(bold=True, size=fontsize)
    else:
        c.font = Font(size=fontsize)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = BORDE
    return c

def hdr(ws, row, cols, bg='2E75B6', fg='FFFFFF'):
    for i, v in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(color=fg, bold=True, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDE

def titulo(ws, text, cols):
    ws.merge_cells(f'A1:{get_column_letter(cols)}1')
    c = ws.cell(row=1, column=1, value=text)
    c.fill = PatternFill("solid", fgColor='1F4E79')
    c.font = Font(color='FFFFFF', bold=True, size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

def generar_excel(pac, adms, inds, facts, logs):
    wb = openpyxl.Workbook()

    # ── Hoja 1: Administrados corregidos ──
    ws1 = wb.active
    ws1.title = 'Administrados'
    titulo(ws1, f'MEDICAMENTOS ADMINISTRADOS — {pac.nombre}', 6)
    hdr(ws1, 2, ['FECHA','ENFERMERO','NOMBRE ORIGINAL (COMERCIAL)','NOMBRE CORREGIDO (FACTURADO)','ESTADO','OBSERVACIÓN'])
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 52
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 30
    for r, a in enumerate(adms, 3):
        fill = ESTADO_FILL.get(a.estado, 'FFFFFF')
        cel(ws1, r, 1, a.fecha, fill, align='center')
        cel(ws1, r, 2, a.hora_enfermero or '', fill)
        cel(ws1, r, 3, a.nombre_original, fill)
        cel(ws1, r, 4, a.nombre_corregido or '— SIN MAPEO —', fill, bold=(a.estado=='ok'))
        cel(ws1, r, 5, a.estado.upper().replace('_',' '), fill, align='center')
        cel(ws1, r, 6, a.observacion or '', fill)
        ws1.row_dimensions[r].height = 20

    # ── Hoja 2: Indicados ──
    ws2 = wb.create_sheet('Indicados')
    titulo(ws2, f'INDICACIONES MÉDICAS — {pac.nombre}', 5)
    hdr(ws2, 2, ['FECHA','MÉDICO','DROGA','DOSIS / VÍA','OBSERVACIÓN'])
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 35
    ws2.column_dimensions['D'].width = 42
    ws2.column_dimensions['E'].width = 30
    for r, i in enumerate(inds, 3):
        cel(ws2, r, 1, i.fecha, align='center')
        cel(ws2, r, 2, i.medico or '')
        cel(ws2, r, 3, i.droga)
        cel(ws2, r, 4, i.dosis_via or '')
        cel(ws2, r, 5, i.observacion or '')
        ws2.row_dimensions[r].height = 18

    # ── Hoja 3: Facturados ──
    ws3 = wb.create_sheet('Facturados')
    titulo(ws3, f'MEDICAMENTOS FACTURADOS (LIQUIDACIÓN) — {pac.nombre}', 5)
    hdr(ws3, 2, ['FECHA','NOMBRE','CANTIDAD','CÓDIGO','VALOR'])
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 16
    for r, f in enumerate(facts, 3):
        cel(ws3, r, 1, f.fecha, align='center')
        cel(ws3, r, 2, f.nombre)
        cel(ws3, r, 3, f.cantidad, align='center')
        cel(ws3, r, 4, f.codigo or '')
        cel(ws3, r, 5, f.valor or '')
        ws3.row_dimensions[r].height = 18

    # ── Hoja 4: Log de cambios ──
    ws4 = wb.create_sheet('Log de Cambios')
    titulo(ws4, f'HISTORIAL DE CAMBIOS — {pac.nombre}', 6)
    hdr(ws4, 2, ['FECHA/HORA','ENTIDAD','CAMPO','VALOR ANTERIOR','VALOR NUEVO','DESCRIPCIÓN'])
    ws4.column_dimensions['A'].width = 18
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 18
    ws4.column_dimensions['D'].width = 35
    ws4.column_dimensions['E'].width = 35
    ws4.column_dimensions['F'].width = 45
    for r, l in enumerate(logs, 3):
        fill = 'E2EFDA' if l.usuario == 'sistema' else 'FFF2CC'
        cel(ws4, r, 1, l.fecha.strftime('%d/%m/%Y %H:%M:%S') if l.fecha else '', fill)
        cel(ws4, r, 2, l.entidad or '', fill, align='center')
        cel(ws4, r, 3, l.campo or '', fill)
        cel(ws4, r, 4, l.valor_anterior or '', fill)
        cel(ws4, r, 5, l.valor_nuevo or '', fill)
        cel(ws4, r, 6, l.descripcion or '', fill)
        ws4.row_dimensions[r].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
